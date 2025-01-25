import pandas as pd
from openai import OpenAI
from config import API_KEY
from openpyxl import load_workbook
import json
import os

class SubjectiveScorer:
    def __init__(self):
        self.client = OpenAI(api_key=API_KEY, base_url="https://api.deepseek.com")
        
    def check_files_exist(self):
        """检查必要文件是否存在"""
        required_files = ["试题与答案.xlsx", "答卷.xlsx"]
        for file in required_files:
            if not os.path.exists(file):
                raise FileNotFoundError(f"缺少必要文件: {file}")
        
    def load_questions(self):
        """加载试题和评分标准"""
        try:
            df = pd.read_excel("试题与答案.xlsx")
            required_columns = ["序号", "试题", "参考答案与评分标准"]
            if not all(col in df.columns for col in required_columns):
                raise ValueError("试题与答案.xlsx缺少必要列")
            return df.to_dict("records")
        except Exception as e:
            print(f"加载试题失败: {str(e)}")
            return []
    
    def load_answers(self):
        """加载学生答卷"""
        try:
            df = pd.read_excel("答卷.xlsx")
            required_columns = ["答题人", "回答内容", "问题序号"]
            if not all(col in df.columns for col in required_columns):
                raise ValueError("答卷.xlsx缺少必要列")
            return df.to_dict("records")
        except Exception as e:
            print(f"加载答卷失败: {str(e)}")
            return []
    
    def score_answer(self, question, answer, criteria):
        """使用deepseek API进行评分"""
        prompt = f"""
        请根据以下评分标准对答案进行评分：
        题目：{question}
        评分标准：{criteria}
        学生答案：{answer}
        
        请返回以下格式的JSON：
        {{
            "score": 分数,
            "analysis": "评分分析"
        }}
        """
        
        try:
            response = self.client.chat.completions.create(
                model="deepseek-chat",
                messages=[
                    {"role": "system", "content": "你是一个专业的评分助手"},
                    {"role": "user", "content": prompt}
                ],
                response_format={"type": "json_object"},
                temperature=0.2
            )
            result = json.loads(response.choices[0].message.content)
            if not all(key in result for key in ["score", "analysis"]):
                raise ValueError("API返回格式不正确")
            return result
        except json.JSONDecodeError:
            print("API返回的JSON解析失败")
            return {"score": 0, "analysis": "评分失败"}
        except Exception as e:
            print(f"API调用失败: {str(e)}")
            return {"score": 0, "analysis": "评分失败"}
    
    def save_results(self, results):
        """保存评分结果"""
        try:
            wb = load_workbook("答卷.xlsx")
            ws = wb.active
            
            for i, result in enumerate(results, start=2):  # 从第二行开始
                ws.cell(row=i, column=3, value=result["score"])
                ws.cell(row=i, column=4, value=result["analysis"])
                
            wb.save("答卷.xlsx")
        except Exception as e:
            print(f"保存结果失败: {str(e)}")
            raise
        
    def run(self):
        """主运行逻辑"""
        try:
            print("开始检查文件...")
            self.check_files_exist()
            print("文件检查通过")
            
            print("加载试题...")
            questions = self.load_questions()
            if not questions:
                print("没有加载到任何试题")
                return
            print(f"成功加载 {len(questions)} 道试题")
                
            print("加载答卷...")
            answers = self.load_answers()
            if not answers:
                print("没有加载到任何答卷")
                return
            print(f"成功加载 {len(answers)} 份答卷")
                
            results = []
            print("开始评分...")
            
            for i, answer in enumerate(answers, 1):
                try:
                    print(f"\n正在处理第 {i}/{len(answers)} 份答卷")
                    print(f"答题人: {answer['答题人']}")
                    print(f"问题序号: {answer['问题序号']}")
                    
                    question = next(q for q in questions if q["序号"] == answer["问题序号"])
                    print("找到对应试题，开始评分...")
                    
                    scoring_result = self.score_answer(
                        question["试题"],
                        answer["回答内容"],
                        question["参考答案与评分标准"]
                    )
                    print(f"评分结果: {scoring_result}")
                    
                    results.append({
                        "序号": answer["问题序号"],
                        "答题人": answer["答题人"],
                        **scoring_result
                    })
                except StopIteration:
                    print(f"找不到问题序号 {answer['问题序号']} 对应的试题")
                    continue
                except Exception as e:
                    print(f"评分过程中出现错误: {str(e)}")
                    continue
                    
            if results:
                print("\n保存评分结果...")
                self.save_results(results)
                print("评分完成，结果已保存到答卷.xlsx")
            else:
                print("没有可保存的评分结果")
                
        except Exception as e:
            print(f"程序运行失败: {str(e)}")

if __name__ == "__main__":
    scorer = SubjectiveScorer()
    scorer.run()
